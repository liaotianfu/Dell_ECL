import re
import os
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from typing import List, Dict, Optional
from html.parser import HTMLParser

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    import openpyxl.styles
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# =========================
# Constants
# =========================
MM_TO_MIL = 39.3700787  # 1 mm = 39.3700787 mils

# =========================
# Helper Functions
# =========================
def get_refdes(pin_name: str) -> Optional[str]:
    if not pin_name:
        return None
    pin_name = pin_name.replace('*', '')
    if pin_name.upper() == 'T':
        return None
    if pin_name.upper().startswith('VIA'):
        return None
    if '.' in pin_name:
        return pin_name.split('.')[0]
    return pin_name

def is_component_pin(name: str) -> bool:
    n = name.replace('*', '').upper()
    if n.startswith('VIA'):
        return False
    if n == 'T':
        return False
    return True

# =========================
# Regex Patterns
# =========================
re_item = re.compile(r'^\s*([A-Za-z0-9_\.\/\*]+)\s+([\-0-9\.]+)\s+([\-0-9\.]+)\s+([LBDV])?\s*([0-9\.]+)\s*(.*)?$')
re_total = re.compile(r'^\s*TOTAL\s+(\d+)\s+VIA\(S\)\s+([0-9\.]+)\s+(mils|millimeters|mm|MILS|MILLIMETERS|MM)', re.IGNORECASE)

def auto_adjust_column_width(ws):
    """自動調整工作表所有欄位寬度"""
    for column_cells in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column_cells:
            if column_letter is None:
                column_letter = cell.column_letter
            try:
                if cell.value:
                    cell_value = str(cell.value)
                    if cell_value.startswith('='):
                        cell_length = 15
                    else:
                        cell_length = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max(max_length + 2, 8), 50)
        if column_letter:
            ws.column_dimensions[column_letter].width = adjusted_width

def add_borders_to_sheet(ws):
    """為工作表中所有有內容的儲存格添加邊框"""
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.value != '':
                cell.border = thin_border

# =========================
# HTML Parser
# =========================
class SimpleTableParser(HTMLParser):
    """Parse first <table> into rows list."""
    def __init__(self):
        super().__init__()
        self.in_table = False
        self.in_tr = False
        self.in_td = False
        self.rows = []
        self.current_row = []
        self._buf = []

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag == "table" and not self.in_table:
            self.in_table = True
        elif self.in_table and tag == "tr":
            self.in_tr = True
            self.current_row = []
        elif self.in_tr and tag in ("td", "th"):
            self.in_td = True
            self._buf = []

    def handle_data(self, data):
        if self.in_td:
            self._buf.append(data)

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag == "table" and self.in_table:
            self.in_table = False
        elif tag == "tr" and self.in_tr:
            self.in_tr = False
            row = [c.strip() for c in self.current_row if c.strip() != ""]
            if row:
                self.rows.append(row)
        elif tag in ("td", "th") and self.in_td:
            self.in_td = False
            self.current_row.append(("".join(self._buf)).strip())

def detect_unit_from_header(header_cells: List[str]) -> str:
    h = " ".join((c or "").lower() for c in header_cells)
    if "mm" in h or "millimeter" in h or "millimeters" in h:
        return "mm"
    return "mils"

def parse_trace_html_to_map(html_path: str) -> Dict[str, List[Dict]]:
    """
    Return: net_name -> list of rows:
      {layer, line_width_mil, len_at_width_mil}
    If HTML is mm, convert to mils.
    """
    with open(html_path, "r", encoding="utf-8", errors="ignore") as f:
        text = f.read()

    p = SimpleTableParser()
    p.feed(text)

    header_idx = None
    for i, row in enumerate(p.rows):
        joined = " ".join(row).lower()
        if "net name" in joined and "layer name" in joined:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("HTML header not found (Net Name / Layer Name).")

    header = [h.strip() for h in p.rows[header_idx]]
    unit = detect_unit_from_header(header)
    scale = MM_TO_MIL if unit == "mm" else 1.0

    def to_float(x):
        try:
            return float(str(x).strip())
        except:
            return None

    out = {}
    for row in p.rows[header_idx + 1:]:
        if len(row) < 5:
            continue
        net = row[0].strip()
        layer = row[1].strip()
        width = to_float(row[3])
        len_at_width = to_float(row[4])
        if width is not None:
            width *= scale
        if len_at_width is not None:
            len_at_width *= scale

        out.setdefault(net, []).append({
            "layer": layer,
            "line_width_mil": width,
            "len_at_width_mil": len_at_width,
        })
    return out

def parse_neckdown_input(s: Optional[str]) -> Optional[float]:
    """
    '4' -> mil
    '4mil'/'4mils' -> mil
    '0.1mm' -> mil
    """
    if s is None:
        return None
    s = s.strip().lower().replace(" ", "")
    if not s:
        return None
    if s.endswith("mm"):
        return float(s[:-2]) * MM_TO_MIL
    if s.endswith("mils"):
        return float(s[:-4])
    if s.endswith("mil"):
        return float(s[:-3])
    return float(s)

def calc_neckdown_layer_map(html_rows: List[Dict], neckdown_width_mil: Optional[float]) -> Dict[str, float]:
    """
    Sum len_at_width by layer where width <= threshold (<= boundary).
    Return: layer -> neckdown_len_mil
    """
    nd = {}
    if neckdown_width_mil is None:
        return nd
    EPS = 1e-9
    for row in html_rows or []:
        lyr = (row.get("layer") or "").strip()
        w = row.get("line_width_mil")
        l = row.get("len_at_width_mil")
        if not lyr or w is None or l is None:
            continue
        if float(w) <= float(neckdown_width_mil) + EPS:  # <= boundary
            nd[lyr] = nd.get(lyr, 0.0) + float(l)
    return nd

# =========================
# PCB loss sheet
# =========================
def create_pcb_loss_sheet(wb):
    if 'PCB loss' in wb.sheetnames:
        ws = wb['PCB loss']
    else:
        ws = wb.create_sheet('PCB loss')

    data_sl = [
        ["Dell's Material Category", "Type", "8 GHz", "16 GHz"],
        ["Mid loss", "SL", -1.16, -2.1],
        ["Low loss 1", "SL", -0.67, -1.15],
        ["Low loss 2", "SL", -0.58, -0.99],
        ["Ultra low loss 1", "SL", -0.54, -0.89],
        ["Ultra low loss 2", "SL", -0.49, -0.79],
        ["Ultra low loss 3", "SL", -0.43, -0.69],
        ["Ultra low loss 4", "SL", -0.4, -0.64],
        ["Mid loss@WC", "SL", -1.45, -2.63],
        ["Low loss 1@WC", "SL", -0.81, -1.38],
        ["Low loss2@WC", "SL", -0.7, -1.19],
        ["Ultra low loss 1@WC", "SL", -0.65, -1.07],
        ["Ultra low loss 2@WC", "SL", -0.59, -0.95],
        ["Ultra low loss 3@WC", "SL", -0.52, -0.83],
        ["Ultra low loss 4@WC", "SL", -0.48, -0.77],
        ["Mid loss@60°C", "SL", -1.35, -2.44],
        ["Low loss 1@60°C", "SL", -0.75, -1.28],
        ["Low loss2@60°C", "SL", -0.65, -1.1],
        ["Ultra low loss 1@60°C", "SL", -0.59, -0.97],
        ["Ultra low loss 2@60°C", "SL", -0.53, -0.86],
        ["Ultra low loss 3@60°C", "SL", -0.47, -0.75],
        ["Ultra low loss 4@60°C", "SL", -0.44, -0.7],
        ["Mid loss@90°C", "SL", -1.47, -2.65],
        ["Low loss 1@90°C", "SL", -0.79, -1.35],
        ["Low loss2@90°C", "SL", -0.68, -1.16],
        ["Ultra low loss 1@90°C", "SL", -0.62, -1.01],
        ["Ultra low loss 2@90°C", "SL", -0.56, -0.9],
        ["Ultra low loss 3@90°C", "SL", -0.49, -0.78],
        ["Ultra low loss 4@90°C", "SL", -0.46, -0.73]
    ]

    data_ms = [
        ["Mid loss", "MS", -1.27, -2.32],
        ["Low loss 1", "MS", -0.81, -1.47],
        ["Low loss 2", "MS", -0.7, -1.26],
        ["Ultra low loss 1", "MS", -0.92, -1.66],
        ["Ultra low loss 2", "MS", -0.82, -1.47],
        ["Ultra low loss 3", "MS", -0.72, -1.29],
        ["Ultra low loss 4", "MS", -0.67, -1.19],
        ["Mid loss@WC", "MS", -1.59, -2.9],
        ["Low loss 1@WC", "MS", -0.98, -1.77],
        ["Low loss2@WC", "MS", -0.84, -1.52],
        ["Ultra low loss 1@WC", "MS", -1.11, -2],
        ["Ultra low loss 2@WC", "MS", -0.99, -1.77],
        ["Ultra low loss 3@WC", "MS", -0.87, -1.55],
        ["Ultra low loss 4@WC", "MS", -0.81, -1.43],
        ["Mid loss@60°C", "MS", -1.48, -2.7],
        ["Low loss 1@60°C", "MS", -0.9, -1.64],
        ["Low loss2@60°C", "MS", -0.78, -1.4],
        ["Ultra low loss 1@60°C", "MS", -1, -1.8],
        ["Ultra low loss 2@60°C", "MS", -0.89, -1.59],
        ["Ultra low loss 3@60°C", "MS", -0.78, -1.4],
        ["Ultra low loss 4@60°C", "MS", -0.73, -1.29],
        ["Mid loss@90°C", "MS", -1.61, -2.93],
        ["Low loss 1@90°C", "MS", -0.95, -1.72],
        ["Low loss2@90°C", "MS", -0.82, -1.48],
        ["Ultra low loss 1@90°C", "MS", -1.04, -1.88],
        ["Ultra low loss 2@90°C", "MS", -0.93, -1.67],
        ["Ultra low loss 3@90°C", "MS", -0.82, -1.46],
        ["Ultra low loss 4@90°C", "MS", -0.76, -1.35]
    ]

    full_data = data_sl + data_ms
    categories = []
    seen = set()

    for row in data_sl[1:]:
        name = row[0]
        if name not in seen:
            categories.append(name)
            seen.add(name)

    for row_idx, row_data in enumerate(full_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.font = openpyxl.styles.Font(bold=True)
                if col_idx >= 3:
                    cell.fill = openpyxl.styles.PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                    cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    for i, cat in enumerate(categories, 1):
        ws.cell(row=i, column=10, value=cat)

    auto_adjust_column_width(ws)
    add_borders_to_sheet(ws)
    return categories

# =========================
# ECL parsing
# =========================
def parse_file(path: str) -> List[Dict]:
    nets = []
    current_net_name = None
    items = []

    with open(path, 'r', encoding='utf-8', errors='ignore') as f:
        lines = [ln.strip() for ln in f]

    for ln in lines:
        if not ln:
            continue
        if ln.startswith('|') or ln.startswith('ECL ') or ln.startswith('Page') or ln.startswith('C:/'):
            continue
        if ln.startswith('dimensions') or ln.startswith('refdes'):
            continue
        if ln.startswith('net name') or ln.startswith('End of ECL') or ln.startswith('total path length'):
            continue
        if ' - - - ' in ln:
            continue
        if ln.startswith('\x0c'):
            continue

        m_total = re_total.match(ln)
        if m_total:
            if current_net_name and items:
                raw_total_len = float(m_total.group(2))
                unit_str = m_total.group(3).lower()
                is_mm = 'mm' in unit_str or 'millimeter' in unit_str
                scale_factor = MM_TO_MIL if is_mm else 1.0

                final_total_len = raw_total_len * scale_factor
                start_pin = items[0][0].replace('*', '')

                converted_items = []
                for (nm, length, lyr, typ) in items:
                    converted_items.append((nm, length * scale_factor, lyr, typ))

                nets.append({
                    'source_file': os.path.basename(path),
                    'net_name': current_net_name,
                    'start_pin': start_pin,
                    'items': converted_items,
                    'total': final_total_len,
                    'via_count_reported': int(m_total.group(1))
                })

            current_net_name = None
            items = []
            continue

        if not re.search(r'[\-0-9]+\.[0-9]+', ln):
            current_net_name = ln.strip()
            items = []
            continue

        m_item = re_item.match(ln)
        if m_item and current_net_name:
            name = m_item.group(1)
            type_char = m_item.group(4)
            length = float(m_item.group(5))
            layer = m_item.group(6) if m_item.group(6) else ''
            items.append((name, length, layer, type_char))

    return nets

def convert_net_to_segments(n: Dict) -> Dict:
    items = n['items']
    if not items:
        return {}

    start_pin = n['start_pin']
    end_pin_name, end_total_len, _, _ = items[-1]
    end_pin = end_pin_name.replace('*', '')

    raw_segments = []
    via_count_calc = 0
    prev_len = 0.0

    last_via_location = None
    VIA_TOLERANCE = 1.0

    for i in range(1, len(items)):
        curr_name, curr_len, curr_layer, _ = items[i]
        curr_name_clean = curr_name.replace('*', '')
        curr_name_upper = curr_name_clean.upper()

        seg_len = curr_len - prev_len
        seg_layer = (curr_layer or '').strip()

        is_via = curr_name_upper.startswith('VIA')
        if is_via:
            if last_via_location is None:
                via_count_calc += 1
                last_via_location = curr_len
            else:
                if abs(curr_len - last_via_location) > VIA_TOLERANCE:
                    via_count_calc += 1
                    last_via_location = curr_len

        next_conn = None
        if i < len(items) - 1:
            if is_via:
                next_conn = 'VIA'
            elif curr_name_upper == 'T':
                next_conn = None
            else:
                next_conn = get_refdes(curr_name_clean)

        raw_segments.append({'layer': seg_layer, 'length': seg_len, 'next_conn': next_conn})
        prev_len = curr_len

    final_segments = []
    for seg in raw_segments:
        if seg['length'] < 0.02:
            if final_segments and seg.get('next_conn'):
                final_segments[-1]['next_conn'] = seg['next_conn']
        else:
            final_segments.append(seg)

    return {
        'source_file': n['source_file'],
        'net_name': n['net_name'],
        'start_pin': start_pin,
        'end_pin': end_pin,
        'segments': final_segments,
        'via_count': via_count_calc,
        'total': end_total_len
    }

def merge_passive_nets(rows: List[Dict]) -> List[Dict]:
    def get_clean_name(name):
        if '_C_' in name:
            return name.replace('_C_', '_')
        if '_R_' in name:
            return name.replace('_R_', '_')
        return name

    while True:
        merged_any = False
        comp_map = {}

        for i, row in enumerate(rows):
            s_pin = row['start_pin']
            e_pin = row['end_pin']
            for pin, ptype in [(s_pin, 'Start'), (e_pin, 'End')]:
                ref = get_refdes(pin)
                if ref and (ref.startswith('C') or ref.startswith('R') or ref.startswith('L')):
                    comp_map.setdefault(ref, []).append((i, pin, ptype))

        indices_to_remove = set()
        new_rows = []

        for ref in sorted(comp_map.keys()):
            entries = comp_map[ref]
            valid_entries = [e for e in entries if e[0] not in indices_to_remove]
            if len(valid_entries) != 2:
                continue

            (idx1, pin1, type1), (idx2, pin2, type2) = valid_entries
            if pin1 == pin2:
                continue

            row1 = rows[idx1]
            row2 = rows[idx2]

            indices_to_remove.add(idx1)
            indices_to_remove.add(idx2)
            merged_any = True

            segs1 = row1['segments'][:]
            start_pin_final = row1['start_pin']
            if type1 == 'Start':
                rev_segs = []
                conns = [s['next_conn'] for s in segs1[:-1]]
                bodies = [{'layer': s['layer'], 'length': s['length']} for s in segs1]
                bodies.reverse()
                conns.reverse()
                for x in range(len(bodies)):
                    s = bodies[x]
                    nxt = conns[x] if x < len(conns) else None
                    s['next_conn'] = nxt
                    rev_segs.append(s)
                segs1 = rev_segs
                start_pin_final = row1['end_pin']

            segs2 = row2['segments'][:]
            end_pin_final = row2['end_pin']
            if type2 == 'End':
                rev_segs = []
                conns = [s['next_conn'] for s in segs2[:-1]]
                bodies = [{'layer': s['layer'], 'length': s['length']} for s in segs2]
                bodies.reverse()
                conns.reverse()
                for x in range(len(bodies)):
                    s = bodies[x]
                    nxt = conns[x] if x < len(conns) else None
                    s['next_conn'] = nxt
                    rev_segs.append(s)
                segs2 = rev_segs
                end_pin_final = row2['start_pin']

            if segs1:
                segs1[-1]['next_conn'] = ref

            merged_segs = segs1 + segs2
            total_len = sum(s['length'] for s in merged_segs)
            new_via_count = row1['via_count'] + row2['via_count']

            n1 = row1['net_name']
            n2 = row2['net_name']
            clean1 = get_clean_name(n1)
            clean2 = get_clean_name(n2)
            final_name = clean2 if len(clean2) < len(n2) else clean1

            new_rows.append({
                'source_file': row1['source_file'],
                'net_name': final_name,
                'start_pin': start_pin_final,
                'end_pin': end_pin_final,
                'segments': merged_segs,
                'via_count': new_via_count,
                'total': total_len
            })

        if not merged_any:
            break

        final_list = [r for i, r in enumerate(rows) if i not in indices_to_remove]
        final_list.extend(new_rows)
        rows = final_list

    return rows

def collect_and_sort_layers(grouped_data: Dict) -> List[str]:
    all_layers = set()
    for (_, _), rows in grouped_data.items():
        for row in rows:
            for seg in row['segments']:
                layer = seg['layer']
                if layer:
                    all_layers.add(layer.strip())

    def layer_sort_key(name):
        u = name.upper()
        if 'TOP' in u:
            return -1000
        if 'BOTTOM' in u or 'BOT' in u:
            return 1000
        nums = re.findall(r'\d+', name)
        if nums:
            return int(nums[0])
        return 0

    return sorted(list(all_layers), key=layer_sort_key)

# =========================
# Main combine_to_excel
# =========================
def combine_to_excel(
    input_files: List[str],
    output_file: str,
    include_source: bool = True,
    html_path: Optional[str] = None,
    neckdown_width_mil: Optional[float] = None
):
    if not HAS_OPENPYXL:
        raise ImportError("The 'openpyxl' library is required for Excel export. Please install it via 'pip install openpyxl'.")

    # ✅ OPTION 1 핵심：只有在 html_path + neckdown_width_mil 都存在時才啟用 ND
    include_nd = bool(html_path and neckdown_width_mil is not None)

    html_map = {}
    if include_nd:
        html_map = parse_trace_html_to_map(html_path)

    all_raw_nets = []
    for fp in input_files:
        if not os.path.exists(fp):
            print(f'[WARN] file not found: {fp}')
            continue
        all_raw_nets.extend(parse_file(fp))

    if not all_raw_nets:
        raise RuntimeError('No net resolved.')

    segment_rows = [convert_net_to_segments(n) for n in all_raw_nets]
    merged_rows = merge_passive_nets(segment_rows)

    grouped_data = {}
    for r in merged_rows:
        s_ref = get_refdes(r['start_pin']) or "Unknown"
        e_ref = get_refdes(r['end_pin']) or "Unknown"
        refs = sorted([s_ref, e_ref])
        key = (refs[0], refs[1])
        grouped_data.setdefault(key, []).append(r)

    sorted_layers = collect_and_sort_layers(grouped_data)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    material_categories = create_pcb_loss_sheet(wb)

    DEFAULT_CONFIG_DATA = {
        'PCIe4': {'freq': 8, 'spec': 28, 'overhead': 3},
        'PCIe5': {'freq': 16, 'spec': 36, 'overhead': 3},
        'PCIe6': {'freq': 16, 'spec': 32, 'overhead': 3},
        'xGMI': {'freq': 18.75, 'spec': 32, 'overhead': 3},
        'UPI': {'freq': 18.75, 'spec': 32, 'overhead': 3},
    }
    INTERFACE_OPTIONS = list(DEFAULT_CONFIG_DATA.keys())
    cfg_tbl = "'Config'!$A$2:$D$100"
    interface_range = "'Config'!$A$2:$A$100"

    FIXED_LOSS_DATA = {'VIA': {'8G': 0.4, '16G': 0.4}, 'Cap': {'8G': 0.5, '16G': 0.5}}

    START_DEVICE_DATA = {
        'Root_Complex_PCIe4': {'8G': 5.0, '16G': "NA"},
        'Root_Complex_PCIe5': {'8G': 5.0, '16G': 9.0},
        'Root_Complex_PCIe6': {'8G': 5.0, '16G': 8.0},
        'AIC_PCIe4': {'8G': 8, '16G': "NA"},
        'AIC_PCIe5': {'8G': 8, '16G': 9.5},
        'AIC_PCIe6': {'8G': 8, '16G': 8.5},
        'OCP_SFF_PCIe4': {'8G': 5.5, '16G': "NA"},
        'OCP_SFF_PCIe5': {'8G': 5.5, '16G': 7.0},
        'OCP_SFF_PCIe6': {'8G': 5.5, '16G': 7.0},
        'EDSFF_PCIe4': {'8G': 5.5, '16G': "NA"},
        'EDSFF_PCIe5': {'8G': 5.5, '16G': 7},
        'EDSFF_PCIe6': {'8G': 5.5, '16G': 6},
        'U2_PCIe4': {'8G': 5.5, '16G': "NA"},
        'U2_PCIe5': {'8G': 5.5, '16G': 6.5},
        'M2_PCIe4': {'8G': 6.5, '16G': "NA"},
        'M2_PCIe5': {'8G': 5.5, '16G': 7.5},
        'CEM_CONN_PCIe4': {'8G': 1.5, '16G': "NA"},
        'CEM_CONN_PCIe5': {'8G': "NA", '16G': 1.5},
        'CEM_CONN_PCIe6': {'8G': "NA", '16G': 0.75},
        'EDSFF_CONN_PCIe4': {'8G': 1, '16G': "NA"},
        'EDSFF_CONN_PCIe5': {'8G': 1, '16G': 1},
        'EDSFF_CONN_PCIe6': {'8G': 0.43, '16G': 0.75},
        'GEN-Z_CONN_PCIe5': {'8G': 0.8, '16G': 1},
        'GZN-Z_CONN_PCIe6': {'8G': 0.4, '16G': 0.75},
        'SFF8639_CONN_PCIe5': {'8G': 0.3, '16G': 1},
        'SFF8639_CONN_PCIe6': {'8G': 0.3, '16G': 0.75},
        'NearStack_CONN_PCIe5': {'8G': 0.28, '16G': 0.45},
        'NearStack_CONN_PCIe6': {'8G': 0.3, '16G': 0.5},
        'MCIO_CONN': {'8G': 0.3, '16G': 0.5},
        'Multi-Trak_CONN': {'8G': 0.3, '16G': 0.5},
    }
    END_DEVICE_DATA = dict(START_DEVICE_DATA)

    CABLE_DATA = {
        'CatA': {'8G': 6.5, '16G': 24},
        'CatC': {'8G': 4.3, '16G': 7.5},
        'CatD': {'8G': 4.3, '16G': 7.5},
        'CatE': {'8G': 4.2, '16G': 6.1},
        'CatF': {'8G': 3.4, '16G': 5.1},
    }

    # =========================
    # Config sheet
    # =========================
    config_ws = wb.create_sheet(title='Config')
    config_ws['A1'] = "Interface_List"
    config_ws['B1'] = "Nyquist frequency (GHz)"
    config_ws['C1'] = "Spec"
    config_ws['D1'] = "Overhead"
    for cell in ['A1', 'B1', 'C1', 'D1']:
        config_ws[cell].font = openpyxl.styles.Font(bold=True)
        config_ws[cell].alignment = openpyxl.styles.Alignment(horizontal='center')
        config_ws[cell].fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for i, opt in enumerate(INTERFACE_OPTIONS, start=2):
        defaults = DEFAULT_CONFIG_DATA[opt]
        config_ws[f'A{i}'] = opt
        config_ws[f'B{i}'] = defaults['freq']
        config_ws[f'C{i}'] = defaults['spec']
        config_ws[f'D{i}'] = defaults['overhead']
    auto_adjust_column_width(config_ws)
    add_borders_to_sheet(config_ws)

    # =========================
    # Summary sheet
    # =========================
    summary_ws = wb.create_sheet(title='Summary', index=0)
    summary_headers = [
        'Interface', 'Sheet Name', 'Component Pair', 'Net Count',
        'Shortest net', 'Shortest length (mil)',
        'Smallest loss Net', 'Smallest Loss at 8G', 'Smallest Loss at 16G',
        'Longest net', 'Longest length (mil)',
        'Highest loss Net', 'Highest Loss at 8G', 'Highest Loss at 16G', 'Risk Level'
    ]
    for idx, header_text in enumerate(summary_headers, start=1):
        cell = summary_ws.cell(row=1, column=idx, value=header_text)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))

    summary_row_idx = 2

    # =========================
    # loss data sheet
    # (仍維持 A:E，若 include_nd=False 則把 D/E 隱藏，避免看到 ND)
    # =========================
    loss_data_ws = wb.create_sheet(title='loss data', index=1)

    loss_data_ws['A1'] = 'Layer'
    loss_data_ws['B1'] = '8G (loss/inch)'
    loss_data_ws['C1'] = '16G (loss/inch)'
    loss_data_ws['D1'] = '8G ND (loss/inch)'
    loss_data_ws['E1'] = '16G ND (loss/inch)'
    loss_data_ws['F1'] = "Select Material:"
    loss_data_ws['G1'] = "Mid loss"
    loss_data_ws['G1'].fill = openpyxl.styles.PatternFill(start_color="ADD8E6", fill_type="solid")

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = loss_data_ws[f'{col}1']
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
    loss_data_ws['F1'].alignment = openpyxl.styles.Alignment(horizontal='right')

    cat_range_str = f"'PCB loss'!$J$1:$J${len(material_categories)}"
    dv_mat = DataValidation(type="list", formula1=cat_range_str, allow_blank=False)
    dv_mat.add(loss_data_ws['G1'])
    loss_data_ws.add_data_validation(dv_mat)

    loss_data_ws['F2'] = "ND Multiplier"
    loss_data_ws['F2'].font = openpyxl.styles.Font(bold=True)
    loss_data_ws['F2'].alignment = openpyxl.styles.Alignment(horizontal='right')
    loss_data_ws['G2'] = 1.10
    loss_data_ws['G2'].fill = openpyxl.styles.PatternFill(start_color="ADD8E6", fill_type="solid")

    loss_data_row = 2
    for layer in sorted_layers:
        loss_data_ws[f'A{loss_data_row}'] = layer

        u_layer = layer.upper()
        is_outer = 'TOP' in u_layer or 'BOTTOM' in u_layer or 'BOT' in u_layer
        mat_type = "MS" if is_outer else "SL"

        f_8g = f'=ABS(SUMIFS(\'PCB loss\'!C:C, \'PCB loss\'!A:A, $G$1, \'PCB loss\'!B:B, "{mat_type}"))'
        f_16g = f'=ABS(SUMIFS(\'PCB loss\'!D:D, \'PCB loss\'!A:A, $G$1, \'PCB loss\'!B:B, "{mat_type}"))'
        loss_data_ws[f'B{loss_data_row}'] = f_8g
        loss_data_ws[f'C{loss_data_row}'] = f_16g

        # ND loss = normal * ND multiplier
        loss_data_ws[f'D{loss_data_row}'] = f"=B{loss_data_row}*$G$2"
        loss_data_ws[f'E{loss_data_row}'] = f"=C{loss_data_row}*$G$2"

        loss_data_row += 1

    for item, vals in FIXED_LOSS_DATA.items():
        loss_data_ws[f'A{loss_data_row}'] = item
        loss_data_ws[f'B{loss_data_row}'] = vals['8G']
        loss_data_ws[f'C{loss_data_row}'] = vals['16G']
        loss_data_ws[f'D{loss_data_row}'] = vals['8G']
        loss_data_ws[f'E{loss_data_row}'] = vals['16G']
        loss_data_row += 1

    # Start Device block
    loss_data_ws['H1'] = "Start Device List"
    loss_data_ws['I1'] = '8G (loss/unit)'
    loss_data_ws['J1'] = '16G (loss/unit)'
    for col in ['H', 'I', 'J']:
        cell = loss_data_ws[f'{col}1']
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    start_device_row = 2
    for device, loss_values in START_DEVICE_DATA.items():
        loss_data_ws[f'H{start_device_row}'] = device
        loss_data_ws[f'I{start_device_row}'] = loss_values['8G']
        loss_data_ws[f'J{start_device_row}'] = loss_values['16G']
        start_device_row += 1
    start_dev_range = "'loss data'!$H$2:$H$100"

    # End Device block
    loss_data_ws['L1'] = "End Device List"
    loss_data_ws['M1'] = '8G (loss/unit)'
    loss_data_ws['N1'] = '16G (loss/unit)'
    for col in ['L', 'M', 'N']:
        cell = loss_data_ws[f'{col}1']
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    end_device_row = 2
    for device, loss_values in END_DEVICE_DATA.items():
        loss_data_ws[f'L{end_device_row}'] = device
        loss_data_ws[f'M{end_device_row}'] = loss_values['8G']
        loss_data_ws[f'N{end_device_row}'] = loss_values['16G']
        end_device_row += 1
    end_dev_range = "'loss data'!$L$2:$L$100"

    # Cable block
    loss_data_ws['P1'] = "Cable List"
    loss_data_ws['Q1'] = '8G (loss/1000mm)'
    loss_data_ws['R1'] = '16G (loss/1000mm)'
    for col in ['P', 'Q', 'R']:
        cell = loss_data_ws[f'{col}1']
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    cable_row = 2
    for cable, loss_values in CABLE_DATA.items():
        loss_data_ws[f'P{cable_row}'] = cable
        loss_data_ws[f'Q{cable_row}'] = loss_values['8G']
        loss_data_ws[f'R{cable_row}'] = loss_values['16G']
        cable_row += 1
    cable_range = "'loss data'!$P$2:$P$100"

    # ✅ OPTION 1：未啟用 ND 時，把 loss data 的 ND 欄位隱藏
    if not include_nd:
        loss_data_ws.column_dimensions['D'].hidden = True
        loss_data_ws.column_dimensions['E'].hidden = True
        # 若你也想把 ND Multiplier 相關提示藏起來，可自行加上隱藏 F2/G2 所在欄或改版面

    auto_adjust_column_width(loss_data_ws)
    add_borders_to_sheet(loss_data_ws)

    def should_have_loss_calc(net_name: str) -> bool:
        name_upper = net_name.strip().upper()
        return name_upper.startswith('PE') or name_upper.startswith('XGMI') or name_upper.startswith('UPI')

    summary_info_list = []

    # =========================
    # Data sheets (✅ include_nd 决定是否插入 ND 欄位)
    # =========================
    for (ref1, ref2), rows in sorted(grouped_data.items()):
        sheet_name = f"{ref1}_{ref2}"
        for ch in '[]:*?/\\':
            sheet_name = sheet_name.replace(ch, '_')
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        ws = wb.create_sheet(title=sheet_name)
        ws['A1'] = "Return to Summary"
        ws['A1'].hyperlink = "#'Summary'!A1"
        ws['A1'].font = openpyxl.styles.Font(color="0563C1", underline="single")

        has_loss_calc_net = any(should_have_loss_calc(r['net_name']) for r in rows)

        max_segments = 0
        for r in rows:
            max_segments = max(max_segments, len(r['segments']))

        header = []
        if include_source:
            header.append('source_file')
        if has_loss_calc_net:
            header.append('Start Device')

        header.append('net_name')
        header.append('Start Pin')

        # ✅ OPTION 1：只在 include_nd=True 時插入 Neck_down 欄位
        for i in range(1, max_segments + 1):
            header += [f'Layer_{i}', f'Length_{i}']
            if include_nd:
                header += [f'Layer_{i}_Neck_down', f'Layer_{i}_Neck_down_Length']
            if i < max_segments:
                header += [f'VIA_{i}']

        header.append('End Pin')

        if has_loss_calc_net:
            header.append('End Device')
            header += ['Cable level', 'Cable length (mm)']

        header += ['total length', 'via count']

        if has_loss_calc_net:
            header += ['Loss at 8G', 'Loss at 16G']

        ws.append(header)

        net_name_idx = header.index('net_name') + 1
        total_len_idx = header.index('total length') + 1
        net_name_col_letter = get_column_letter(net_name_idx)
        total_len_col_letter = get_column_letter(total_len_idx)

        if has_loss_calc_net:
            loss8g_idx = header.index('Loss at 8G') + 1
            loss16g_idx = header.index('Loss at 16G') + 1
            loss8g_col = get_column_letter(loss8g_idx)
            loss16g_col = get_column_letter(loss16g_idx)
        else:
            loss8g_col = loss16g_col = None

        # Header style row=2
        for col_idx in range(1, len(header) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        rows.sort(key=lambda x: x['net_name'])
        current_row_idx = 3

        # Pre-calc per segment col indices
        layer_col_idx = [header.index(f'Layer_{k}') + 1 for k in range(1, max_segments + 1)]
        len_col_idx = [header.index(f'Length_{k}') + 1 for k in range(1, max_segments + 1)]
        if include_nd:
            ndlen_col_idx = [header.index(f'Layer_{k}_Neck_down_Length') + 1 for k in range(1, max_segments + 1)]
        else:
            ndlen_col_idx = None

        for r in rows:
            row_data = []
            if include_source:
                row_data.append(r['source_file'])

            start_device_col_idx = -1
            if has_loss_calc_net:
                start_device_col_idx = len(row_data)
                row_data.append('Root_Complex_PCIe6')

            row_data += [r['net_name'], r['start_pin']]
            segs = r['segments']

            formula_components_8g = []
            formula_components_16g = []
            is_loss_net = should_have_loss_calc(r['net_name'])

            nd_remaining = {}
            if include_nd:
                nd_remaining = calc_neckdown_layer_map(html_map.get(r['net_name'], []), neckdown_width_mil)

            for i in range(max_segments):
                if i < len(segs):
                    s = segs[i]
                    lyr = (s['layer'] or '').strip()
                    seg_len = float(s['length'])

                    if include_nd:
                        nd_avail = float(nd_remaining.get(lyr, 0.0)) if lyr else 0.0
                        nd_take = min(nd_avail, seg_len) if lyr else 0.0
                        if lyr:
                            nd_remaining[lyr] = nd_avail - nd_take
                        normal_len = seg_len - nd_take

                        row_data += [
                            lyr,
                            round(normal_len, 2),
                            (lyr if nd_take > 0 else ''),
                            (round(nd_take, 2) if nd_take > 0 else '')
                        ]
                    else:
                        # ✅ 未啟用 ND：只填 Layer/Length
                        row_data += [lyr, round(seg_len, 2)]

                    # Loss formula
                    if has_loss_calc_net and is_loss_net and lyr:
                        lyr_col_letter = get_column_letter(layer_col_idx[i])
                        nlen_col_letter = get_column_letter(len_col_idx[i])

                        term_8g_normal = f"(({nlen_col_letter}{current_row_idx}/1000)*VLOOKUP({lyr_col_letter}{current_row_idx},'loss data'!$A:$E,2,0))"
                        term_16g_normal = f"(({nlen_col_letter}{current_row_idx}/1000)*VLOOKUP({lyr_col_letter}{current_row_idx},'loss data'!$A:$E,3,0))"

                        if include_nd:
                            ndlen_col_letter = get_column_letter(ndlen_col_idx[i])
                            term_8g_nd = f"((IF({ndlen_col_letter}{current_row_idx}=\"\",0,{ndlen_col_letter}{current_row_idx})/1000)*VLOOKUP({lyr_col_letter}{current_row_idx},'loss data'!$A:$E,4,0))"
                            term_16g_nd = f"((IF({ndlen_col_letter}{current_row_idx}=\"\",0,{ndlen_col_letter}{current_row_idx})/1000)*VLOOKUP({lyr_col_letter}{current_row_idx},'loss data'!$A:$E,5,0))"
                            formula_components_8g.append(f"({term_8g_normal}+{term_8g_nd})")
                            formula_components_16g.append(f"({term_16g_normal}+{term_16g_nd})")
                        else:
                            # ✅ 未啟用 ND：只用 normal
                            formula_components_8g.append(f"({term_8g_normal})")
                            formula_components_16g.append(f"({term_16g_normal})")

                    if i < max_segments - 1:
                        conn = s.get('next_conn')
                        row_data += [conn if conn else '']
                else:
                    # empty segment
                    if include_nd:
                        row_data += ['', '', '', '']
                    else:
                        row_data += ['', '']
                    if i < max_segments - 1:
                        row_data += ['']

            row_data.append(r['end_pin'])

            end_device_col_idx = -1
            cable_level_col_idx = -1
            cable_length_col_idx = -1
            if has_loss_calc_net:
                end_device_col_idx = len(row_data)
                row_data.append('CEM_CONN_PCIe6')
                cable_level_col_idx = len(row_data)
                cable_length_col_idx = len(row_data) + 1
                row_data += ['CatF', 0]

            row_data += [round(r['total'], 2), r['via_count']]

            if has_loss_calc_net:
                if is_loss_net:
                    cap_cnt = 0
                    for s in segs:
                        if s.get('next_conn') and str(s.get('next_conn')).startswith('C'):
                            cap_cnt += 1

                    start_device_cell = get_column_letter(start_device_col_idx + 1) if start_device_col_idx != -1 else 'B'
                    end_device_cell = get_column_letter(end_device_col_idx + 1) if end_device_col_idx != -1 else 'C'

                    f_start_8g = f"(VLOOKUP({start_device_cell}{current_row_idx},'loss data'!$H:$J,2,0))"
                    f_start_16g = f"(VLOOKUP({start_device_cell}{current_row_idx},'loss data'!$H:$J,3,0))"

                    f_end_8g = f"(VLOOKUP({end_device_cell}{current_row_idx},'loss data'!$L:$N,2,0))"
                    f_end_16g = f"(VLOOKUP({end_device_cell}{current_row_idx},'loss data'!$L:$N,3,0))"

                    if cable_level_col_idx != -1 and cable_length_col_idx != -1:
                        cable_level_cell = get_column_letter(cable_level_col_idx + 1)
                        cable_length_cell = get_column_letter(cable_length_col_idx + 1)
                        f_cable_8g = f"(({cable_length_cell}{current_row_idx}/1000)*VLOOKUP({cable_level_cell}{current_row_idx},'loss data'!$P:$R,2,0))"
                        f_cable_16g = f"(({cable_length_cell}{current_row_idx}/1000)*VLOOKUP({cable_level_cell}{current_row_idx},'loss data'!$P:$R,3,0))"
                    else:
                        f_cable_8g = "0"
                        f_cable_16g = "0"

                    via_count_col = header.index('via count') + 1
                    via_count_cell = get_column_letter(via_count_col)
                    f_via_8g = f"({via_count_cell}{current_row_idx}*VLOOKUP(\"VIA\",'loss data'!$A:$E,2,0))"
                    f_via_16g = f"({via_count_cell}{current_row_idx}*VLOOKUP(\"VIA\",'loss data'!$A:$E,3,0))"

                    if cap_cnt > 0:
                        f_cap_8g = f"({cap_cnt}*VLOOKUP(\"Cap\",'loss data'!$A:$E,2,0))"
                        f_cap_16g = f"({cap_cnt}*VLOOKUP(\"Cap\",'loss data'!$A:$E,3,0))"
                    else:
                        f_cap_8g = "0"
                        f_cap_16g = "0"

                    seg_sum_8g = "+".join(formula_components_8g) if formula_components_8g else "0"
                    seg_sum_16g = "+".join(formula_components_16g) if formula_components_16g else "0"

                    final_f_8g = f"=({f_start_8g})+({seg_sum_8g})+({f_end_8g})+({f_cable_8g})+({f_via_8g})+({f_cap_8g})"
                    final_f_16g = f"=({f_start_16g})+({seg_sum_16g})+({f_end_16g})+({f_cable_16g})+({f_via_16g})+({f_cap_16g})"

                    row_data += [final_f_8g, final_f_16g]
                else:
                    row_data += ['', '']

            ws.append(row_data)

            # Data validations
            if has_loss_calc_net:
                if start_device_col_idx != -1:
                    start_device_cell_obj = ws[f'{get_column_letter(start_device_col_idx + 1)}{current_row_idx}']
                    start_device_cell_obj.fill = openpyxl.styles.PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                    dv_start = DataValidation(type="list", formula1=start_dev_range, allow_blank=False)
                    dv_start.add(start_device_cell_obj)
                    ws.add_data_validation(dv_start)

                if end_device_col_idx != -1:
                    end_device_cell_obj = ws[f'{get_column_letter(end_device_col_idx + 1)}{current_row_idx}']
                    end_device_cell_obj.fill = openpyxl.styles.PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                    dv_end = DataValidation(type="list", formula1=end_dev_range, allow_blank=False)
                    dv_end.add(end_device_cell_obj)
                    ws.add_data_validation(dv_end)

                if cable_level_col_idx != -1:
                    cable_level_cell_obj = ws[f'{get_column_letter(cable_level_col_idx + 1)}{current_row_idx}']
                    cable_level_cell_obj.fill = openpyxl.styles.PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                    dv_cable = DataValidation(type="list", formula1=cable_range, allow_blank=True)
                    dv_cable.add(cable_level_cell_obj)
                    ws.add_data_validation(dv_cable)

            current_row_idx += 1

        auto_adjust_column_width(ws)
        add_borders_to_sheet(ws)

        summary_info_list.append({
            'sheet_name': sheet_name,
            'ref1': ref1,
            'ref2': ref2,
            'net_count': len(rows),
            'net_col': net_name_col_letter,
            'len_col': total_len_col_letter,
            'has_loss': has_loss_calc_net,
            'loss8g_col': loss8g_col,
            'loss16g_col': loss16g_col
        })

    # =========================
    # Fill Summary
    # =========================
    for info in summary_info_list:
        sh_name = info['sheet_name']
        r = summary_row_idx

        summary_ws[f'A{r}'] = 'PCIe6'
        summary_ws[f'A{r}'].fill = openpyxl.styles.PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        dv_interface = DataValidation(type="list", formula1=interface_range, allow_blank=False)
        dv_interface.add(summary_ws[f'A{r}'])
        summary_ws.add_data_validation(dv_interface)

        cell = summary_ws[f'B{r}']
        cell.value = sh_name
        cell.hyperlink = f"#'{sh_name}'!A1"
        cell.font = openpyxl.styles.Font(color="0563C1", underline="single")

        summary_ws[f'C{r}'] = f"{info['ref1']} <-> {info['ref2']}"
        summary_ws[f'D{r}'] = info['net_count']

        len_range_str = f"'{sh_name}'!{info['len_col']}:{info['len_col']}"
        name_range_str = f"'{sh_name}'!{info['net_col']}:{info['net_col']}"

        match_min_part = f"MATCH(MIN({len_range_str}),{len_range_str},0)"
        link_addr_min = f"\"#'{sh_name}'!A\"&{match_min_part}"
        display_name_min = f"INDEX({name_range_str},{match_min_part})"
        summary_ws[f'E{r}'].value = f'=HYPERLINK({link_addr_min}, {display_name_min})'
        summary_ws[f'E{r}'].font = openpyxl.styles.Font(color="0563C1", underline="single")
        summary_ws[f'F{r}'].value = f'=MIN({len_range_str})'
        summary_ws[f'F{r}'].number_format = '0.00'

        if info['has_loss'] and info['loss8g_col'] and info['loss16g_col']:
            loss8g_range = f"'{sh_name}'!{info['loss8g_col']}:{info['loss8g_col']}"
            loss16g_range = f"'{sh_name}'!{info['loss16g_col']}:{info['loss16g_col']}"
            v_freq = f"VLOOKUP($A{r},{cfg_tbl},2,0)"

            match_8g_min = f"MATCH(MIN({loss8g_range}),{loss8g_range},0)"
            match_16g_min = f"MATCH(MIN({loss16g_range}),{loss16g_range},0)"
            link_8g_min = f"\"#'{sh_name}'!A\"&({match_8g_min})"
            link_16g_min = f"\"#'{sh_name}'!A\"&({match_16g_min})"
            f_link_min = f'IF({v_freq}<=8, HYPERLINK({link_8g_min}, INDEX({name_range_str},{match_8g_min})), HYPERLINK({link_16g_min}, INDEX({name_range_str},{match_16g_min})))'
            summary_ws[f'G{r}'].value = f'={f_link_min}'
            summary_ws[f'G{r}'].font = openpyxl.styles.Font(color="0563C1", underline="single")
            summary_ws[f'H{r}'].value = f'=MIN({loss8g_range})'
            summary_ws[f'I{r}'].value = f'=MIN({loss16g_range})'
            summary_ws[f'H{r}'].number_format = '0.00'
            summary_ws[f'I{r}'].number_format = '0.00'

            match_8g_max = f"MATCH(MAX({loss8g_range}),{loss8g_range},0)"
            match_16g_max = f"MATCH(MAX({loss16g_range}),{loss16g_range},0)"
            link_8g_max = f"\"#'{sh_name}'!A\"&({match_8g_max})"
            link_16g_max = f"\"#'{sh_name}'!A\"&({match_16g_max})"
            f_link_max = f'IF({v_freq}<=8, HYPERLINK({link_8g_max}, INDEX({name_range_str},{match_8g_max})), HYPERLINK({link_16g_max}, INDEX({name_range_str},{match_16g_max})))'
            summary_ws[f'L{r}'].value = f'={f_link_max}'
            summary_ws[f'L{r}'].font = openpyxl.styles.Font(color="0563C1", underline="single")
            summary_ws[f'M{r}'].value = f'=MAX({loss8g_range})'
            summary_ws[f'N{r}'].value = f'=MAX({loss16g_range})'
            summary_ws[f'M{r}'].number_format = '0.00'
            summary_ws[f'N{r}'].number_format = '0.00'

            v_spec = f"VLOOKUP($A{r},{cfg_tbl},3,0)"
            v_over = f"VLOOKUP($A{r},{cfg_tbl},4,0)"
            limit_val = f"({v_spec}-{v_over})"
            m_ref = f"$M{r}"
            n_ref = f"$N{r}"
            risk_formula = (
                f'=IFERROR('
                f'IF({v_freq}<=8,'
                f'  IF(ISNUMBER({m_ref}), IF({m_ref}>{limit_val}, "High Risk", "Low Risk"), "No Data"),'
                f'  IF(ISNUMBER({n_ref}), IF({n_ref}>{limit_val}, "High Risk", "Low Risk"), "No Data")'
                f'), "Data Error")'
            )
            summary_ws[f'O{r}'].value = risk_formula
            summary_ws[f'O{r}'].font = openpyxl.styles.Font(bold=True)

        match_max_part = f"MATCH(MAX({len_range_str}),{len_range_str},0)"
        link_addr_max = f"\"#'{sh_name}'!A\"&{match_max_part}"
        display_name_max = f"INDEX({name_range_str},{match_max_part})"
        summary_ws[f'J{r}'].value = f'=HYPERLINK({link_addr_max}, {display_name_max})'
        summary_ws[f'J{r}'].font = openpyxl.styles.Font(color="0563C1", underline="single")
        summary_ws[f'K{r}'].value = f'=MAX({len_range_str})'
        summary_ws[f'K{r}'].number_format = '0.00'

        summary_row_idx += 1

    auto_adjust_column_width(summary_ws)
    add_borders_to_sheet(summary_ws)

    wb.save(output_file)

# =========================
# GUI (Single Window)
# =========================
def run_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox

    if not HAS_OPENPYXL:
        messagebox.showerror(
            'Missing Library',
            "Please install 'openpyxl' to use Excel export feature.\nRun: pip install openpyxl"
        )
        return

    root = tk.Tk()
    root.title("ECL Report Combiner")
    root.geometry("820x520")
    root.minsize(820, 520)

    # -------------------------
    # State
    # -------------------------
    ecl_files = []  # list of paths
    html_path_var = tk.StringVar(value="")
    neck_var = tk.StringVar(value="")     # e.g. 4mil / 0.1mm / blank
    out_path_var = tk.StringVar(value="ecl_reports_combined.xlsx")

    # -------------------------
    # Helpers
    # -------------------------
    def refresh_listbox():
        lb.delete(0, tk.END)
        for p in ecl_files:
            lb.insert(tk.END, p)

    def add_ecl_files():
        paths = filedialog.askopenfilenames(
            title="Choose ECL report files",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        if not paths:
            return

        # avoid duplicates while preserving order
        existing = set(ecl_files)
        for p in paths:
            if p not in existing:
                ecl_files.append(p)
                existing.add(p)
        refresh_listbox()

    def remove_selected():
        sel = list(lb.curselection())
        if not sel:
            return
        for idx in reversed(sel):
            del ecl_files[idx]
        refresh_listbox()

    def clear_all():
        ecl_files.clear()
        refresh_listbox()

    def choose_html():
        p = filedialog.askopenfilename(
            title="Choose HTML trace report (optional)",
            filetypes=[("HTML files", "*.htm *.html"), ("All files", "*.*")]
        )
        if p:
            html_path_var.set(p)

    def clear_html():
        html_path_var.set("")

    def choose_output():
        p = filedialog.asksaveasfilename(
            title="Save output Excel file",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=os.path.basename(out_path_var.get()) if out_path_var.get() else "ecl_reports_combined.xlsx"
        )
        if p:
            out_path_var.set(p)

    def run_job():
        if not ecl_files:
            messagebox.showwarning("No input", "Please add at least one ECL (.txt) file.")
            return

        out_path = out_path_var.get().strip()
        if not out_path:
            messagebox.showwarning("No output", "Please choose an output .xlsx file.")
            return

        html_file = html_path_var.get().strip() or None
        neck_str = neck_var.get().strip()
        neckdown_width_mil = None

        if neck_str:
            try:
                neckdown_width_mil = parse_neckdown_input(neck_str)
            except Exception:
                messagebox.showerror("Invalid neckdown",
                                     "Neckdown format error.\nExamples: 4mil, 4mils, 0.1mm, 4")
                return

        # If threshold is given but no HTML => skip ND (Option 1 behavior)
        if neckdown_width_mil is not None and not html_file:
            messagebox.showinfo("Neckdown skipped",
                                "Neckdown threshold was provided but HTML trace report is not selected.\n"
                                "Neckdown will be skipped (no Neck_down columns).")
            neckdown_width_mil = None

        btn_run.config(state="disabled")
        status_var.set("Running... please wait.")
        root.update_idletasks()

        try:
            combine_to_excel(
                input_files=list(ecl_files),
                output_file=out_path,
                include_source=True,
                html_path=html_file,
                neckdown_width_mil=neckdown_width_mil
            )
            status_var.set("Completed.")
            messagebox.showinfo("Completed", f"Exported:\n{out_path}")
        except Exception as e:
            status_var.set("Failed.")
            messagebox.showerror("Error", f"Processing failure:\n{e}")
            import traceback
            traceback.print_exc()
        finally:
            btn_run.config(state="normal")

    # -------------------------
    # Layout
    # -------------------------
    frm = tk.Frame(root, padx=12, pady=12)
    frm.pack(fill="both", expand=True)

    tk.Label(frm, text="ECL Report Combiner", font=("Segoe UI", 15, "bold")).pack(anchor="w")

    ecl_frame = tk.LabelFrame(frm, text="Input ECL report files (.txt)", padx=10, pady=10)
    ecl_frame.pack(fill="both", expand=True, pady=(10, 10))

    list_frame = tk.Frame(ecl_frame)
    list_frame.pack(fill="both", expand=True)

    sb = tk.Scrollbar(list_frame, orient="vertical")
    lb = tk.Listbox(list_frame, selectmode=tk.EXTENDED, yscrollcommand=sb.set)
    sb.config(command=lb.yview)
    lb.pack(side="left", fill="both", expand=True)
    sb.pack(side="right", fill="y")

    btns = tk.Frame(ecl_frame)
    btns.pack(fill="x", pady=(8, 0))

    tk.Button(btns, text="Add files...", width=12, command=add_ecl_files).pack(side="left")
    tk.Button(btns, text="Remove selected", width=14, command=remove_selected).pack(side="left", padx=(8, 0))
    tk.Button(btns, text="Clear", width=10, command=clear_all).pack(side="left", padx=(8, 0))

    opt_frame = tk.LabelFrame(frm, text="Options", padx=10, pady=10)
    opt_frame.pack(fill="x", pady=(0, 10))

    tk.Label(opt_frame, text="HTML trace report (optional):").grid(row=0, column=0, sticky="w")
    tk.Entry(opt_frame, textvariable=html_path_var).grid(row=0, column=1, sticky="we", padx=(8, 8))
    tk.Button(opt_frame, text="Browse...", command=choose_html).grid(row=0, column=2, sticky="e")
    tk.Button(opt_frame, text="Clear", command=clear_html).grid(row=0, column=3, sticky="e", padx=(8, 0))

    tk.Label(opt_frame, text="Neckdown threshold (blank = skip):").grid(row=1, column=0, sticky="w", pady=(8, 0))
    tk.Entry(opt_frame, textvariable=neck_var).grid(row=1, column=1, sticky="we", padx=(8, 8), pady=(8, 0))
    tk.Label(opt_frame, text="Examples: 4mil, 0.1mm").grid(row=1, column=2, columnspan=2, sticky="w", pady=(8, 0))

    opt_frame.columnconfigure(1, weight=1)

    out_frame = tk.LabelFrame(frm, text="Output", padx=10, pady=10)
    out_frame.pack(fill="x")

    tk.Label(out_frame, text="Output Excel file (.xlsx):").grid(row=0, column=0, sticky="w")
    tk.Entry(out_frame, textvariable=out_path_var).grid(row=0, column=1, sticky="we", padx=(8, 8))
    tk.Button(out_frame, text="Browse...", command=choose_output).grid(row=0, column=2, sticky="e")
    out_frame.columnconfigure(1, weight=1)

    bottom = tk.Frame(frm)
    bottom.pack(fill="x", pady=(10, 0))

    status_var = tk.StringVar(value="Ready.")
    tk.Label(bottom, textvariable=status_var, fg="#444").pack(side="left")

    btn_run = tk.Button(bottom, text="Run", width=12, command=run_job)
    btn_run.pack(side="right")

    root.mainloop()


if __name__ == '__main__':
    try:
        run_gui()
    except tk.TclError:
        print('GUI not available.')