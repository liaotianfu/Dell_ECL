import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import threading
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ==========================================
# Core Logic Functions
# ==========================================

def auto_adjust_column_width(ws):
    """
    自動調整工作表所有欄位寬度
    """
    from openpyxl.cell.cell import MergedCell

    for column in ws.columns:
        max_length = 0
        column_letter = None

        for cell in column:
            # 跳過 MergedCell
            if isinstance(cell, MergedCell):
                continue

            if column_letter is None:
                column_letter = cell.column_letter

            try:
                if cell.value:
                    cell_value = str(cell.value)
                    # 處理公式的情況，使用預估長度
                    if cell_value.startswith('='):
                        cell_length = 15
                    else:
                        # 計算實際文字長度（中文字算2個字元）
                        cell_length = sum(2 if ord(c) > 127 else 1 for c in cell_value)

                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        # 設定欄寬，最小8，最大50
        adjusted_width = min(max(max_length + 2, 8), 50)
        if column_letter:
            ws.column_dimensions[column_letter].width = adjusted_width


def add_borders_to_sheet(ws):
    """
    為工作表中所有有內容的儲存格添加邊框
    """
    from openpyxl.cell.cell import MergedCell

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows():
        for cell in row:
            # 跳過 MergedCell
            if isinstance(cell, MergedCell):
                continue

            if cell.value is not None and cell.value != '':
                cell.border = thin_border


def load_mapping_table(file_path, log_func):
    """Load Mapping Table Excel"""
    log_func(f"Loading Mapping Table: {os.path.basename(file_path)}")
    try:
        df = pd.read_excel(file_path, header=0)

        if df.shape[1] < 2:
            raise ValueError("Mapping Table format error: at least 2 columns required")

        comp_map = defaultdict(list)
        pin_map = {}

        col1 = df.columns[0]
        col2 = df.columns[1]

        for _, row in df.iterrows():
            src = str(row[col1]).strip()
            dst = str(row[col2]).strip()

            if src != 'nan' and dst != 'nan':
                if '.' in src:
                    pin_map[src] = dst
                else:
                    comp_map[src].append(dst)

        log_func(f"Mapping rules: {len(comp_map)} component pairs, {len(pin_map)} specific pin pairs")
        return comp_map, pin_map
    except Exception as e:
        raise ValueError(f"Failed to load Mapping Table: {e}")


def parse_pin(pin_str):
    if pd.isna(pin_str): return None, None
    s = str(pin_str).strip()
    if '.' in s:
        parts = s.split('.', 1)
        return parts[0], parts[1]
    return None, None


def load_all_sheets(file_path, log_func):
    log_func(f"Loading data: {os.path.basename(file_path)}")
    try:
        xls = pd.ExcelFile(file_path)
    except Exception as e:
        raise ValueError(f"Cannot open Excel: {e}")

    all_data = []
    exclude = ['Summary', 'Config', 'loss data']

    for sheet_name in xls.sheet_names:
        if sheet_name in exclude:
            continue
        try:
            df = pd.read_excel(xls, sheet_name=sheet_name, header=1)
            df.columns = df.columns.str.strip()

            if 'Start Pin' not in df.columns:
                continue

            df['Source_Sheet'] = sheet_name
            all_data.append(df)
        except:
            pass

    if not all_data:
        raise ValueError("No valid data sheets found")
    return pd.concat(all_data, ignore_index=True)


def create_config_sheet(wb):
    if 'Config' in wb.sheetnames:
        wb.remove(wb['Config'])
    ws = wb.create_sheet('Config', 1)

    headers = ["Interface_List", "Nyquist frequency (GHz)", "Spec", "Overhead"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    data = [
        ['PCIe4', 8, 28, 3],
        ['PCIe5', 16, 36, 3],
        ['PCIe6', 16, 32, 3],
        ['xGMI', 18.75, 32, 3],
        ['UPI', 18.75, 32, 3]
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 自動調整欄寬並添加邊框
    auto_adjust_column_width(ws)
    add_borders_to_sheet(ws)


def create_summary_sheet(wb, summary_info_list):
    if 'Summary' in wb.sheetnames:
        wb.remove(wb['Summary'])
    ws = wb.create_sheet('Summary', 0)

    # Updated header order + new Smallest loss columns
    headers = [
        'Interface', 'Sheet Name', 'Component Pair', 'Net Count',
        'Shortest net', 'Shortest length (mil)',
        'Smallest loss Net', 'Smallest Loss at 8G', 'Smallest Loss at 16G',
        'Longest net', 'Longest length (mil)',
        'Highest loss Net', 'Highest Loss at 8G', 'Highest Loss at 16G',
        'Risk Level'
    ]

    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(bottom=Side(style='thin'))

    interface_range = "'Config'!$A$2:$A$6"
    dv_interface = DataValidation(type="list", formula1=interface_range, allow_blank=False)
    config_range = "'Config'!$A$2:$D$100"

    for idx, info in enumerate(summary_info_list, 2):
        ws[f'A{idx}'] = 'PCIe6'
        ws[f'A{idx}'].fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        dv_interface.add(ws[f'A{idx}'])

        sheet_name = info['sheet_name']
        ws[f'B{idx}'] = sheet_name
        ws[f'B{idx}'].hyperlink = f"#'{sheet_name}'!A1"
        ws[f'B{idx}'].font = Font(color="0563C1", underline="single")

        ws[f'C{idx}'] = f"{info['ref1']} <-> {info['ref2']}"
        ws[f'D{idx}'] = info['net_count']

        len_col = info.get('total_len_col', 'Z')
        net_col = info.get('net_name_col', 'B')
        loss8g_col = info.get('total_loss8g_col')
        loss16g_col = info.get('total_loss16g_col')
        max_row = info['max_row']

        len_range = f"'{sheet_name}'!${len_col}$3:${len_col}${max_row}"
        net_range = f"'{sheet_name}'!${net_col}$3:${net_col}${max_row}"

        # --- Shortest net (E, F) ---
        match_min_len = f"MATCH(MIN({len_range}),{len_range},0)"
        link_min_len = f"\"#'{sheet_name}'!A\"&({match_min_len}+2)"
        ws[f'E{idx}'].value = f'=HYPERLINK({link_min_len}, INDEX({net_range},{match_min_len}))'
        ws[f'E{idx}'].font = Font(color="0563C1", underline="single")
        ws[f'F{idx}'].value = f'=MIN({len_range})'

        # --- Longest net (J, K) ---
        match_max_len = f"MATCH(MAX({len_range}),{len_range},0)"
        link_max_len = f"\"#'{sheet_name}'!A\"&({match_max_len}+2)"
        ws[f'J{idx}'].value = f'=HYPERLINK({link_max_len}, INDEX({net_range},{match_max_len}))'
        ws[f'J{idx}'].font = Font(color="0563C1", underline="single")
        ws[f'K{idx}'].value = f'=MAX({len_range})'

        if info.get('has_loss'):
            v_freq = f"VLOOKUP($A{idx},{config_range},2,0)"

            rng_8g = f"'{sheet_name}'!${loss8g_col}$3:${loss8g_col}${max_row}" if loss8g_col else None
            rng_16g = f"'{sheet_name}'!${loss16g_col}$3:${loss16g_col}${max_row}" if loss16g_col else None

            # ----------------------
            # Smallest loss (G, H, I) => MIN
            # ----------------------
            if loss8g_col and loss16g_col:
                match_8g_min = f"MATCH(MIN({rng_8g}),{rng_8g},0)"
                match_16g_min = f"MATCH(MIN({rng_16g}),{rng_16g},0)"

                link_8g_min = f"\"#'{sheet_name}'!A\"&({match_8g_min}+2)"
                link_16g_min = f"\"#'{sheet_name}'!A\"&({match_16g_min}+2)"

                min_link_formula = (
                    f'IF({v_freq}<=8, '
                    f'  HYPERLINK({link_8g_min}, INDEX({net_range},{match_8g_min})), '
                    f'  HYPERLINK({link_16g_min}, INDEX({net_range},{match_16g_min}))'
                    f')'
                )
                ws[f'G{idx}'].value = f'={min_link_formula}'
                ws[f'G{idx}'].font = Font(color="0563C1", underline="single")

                row_idx_min = f"IF({v_freq}<=8, {match_8g_min}, {match_16g_min})"
                ws[f'H{idx}'].value = f'=INDEX({rng_8g}, {row_idx_min})'
                ws[f'I{idx}'].value = f'=INDEX({rng_16g}, {row_idx_min})'

            elif loss8g_col:
                match_8g_min = f"MATCH(MIN({rng_8g}),{rng_8g},0)"
                link_8g_min = f"\"#'{sheet_name}'!A\"&({match_8g_min}+2)"

                safe_formula = f'IF({v_freq}<=8, HYPERLINK({link_8g_min}, INDEX({net_range},{match_8g_min})), "No Data")'
                ws[f'G{idx}'].value = f'={safe_formula}'
                ws[f'G{idx}'].font = Font(color="0563C1", underline="single")
                ws[f'H{idx}'].value = f'=INDEX({rng_8g}, {match_8g_min})'
                ws[f'I{idx}'].value = "N/A"

            elif loss16g_col:
                match_16g_min = f"MATCH(MIN({rng_16g}),{rng_16g},0)"
                link_16g_min = f"\"#'{sheet_name}'!A\"&({match_16g_min}+2)"

                safe_formula = f'IF({v_freq}>8, HYPERLINK({link_16g_min}, INDEX({net_range},{match_16g_min})), "No Data")'
                ws[f'G{idx}'].value = f'={safe_formula}'
                ws[f'G{idx}'].font = Font(color="0563C1", underline="single")
                ws[f'H{idx}'].value = "N/A"
                ws[f'I{idx}'].value = f'=INDEX({rng_16g}, {match_16g_min})'

            # ----------------------
            # Highest loss (L, M, N) => MAX (same logic as before, shifted)
            # ----------------------
            if loss8g_col and loss16g_col:
                match_8g = f"MATCH(MAX({rng_8g}),{rng_8g},0)"
                match_16g = f"MATCH(MAX({rng_16g}),{rng_16g},0)"

                link_8g = f"\"#'{sheet_name}'!A\"&({match_8g}+2)"
                link_16g = f"\"#'{sheet_name}'!A\"&({match_16g}+2)"

                full_link_formula = f'IF({v_freq}<=8, HYPERLINK({link_8g}, INDEX({net_range},{match_8g})), HYPERLINK({link_16g}, INDEX({net_range},{match_16g})))'
                ws[f'L{idx}'].value = f'={full_link_formula}'
                ws[f'L{idx}'].font = Font(color="0563C1", underline="single")

                row_idx_calc = f"IF({v_freq}<=8, {match_8g}, {match_16g})"
                ws[f'M{idx}'].value = f'=INDEX({rng_8g}, {row_idx_calc})'
                ws[f'N{idx}'].value = f'=INDEX({rng_16g}, {row_idx_calc})'

            elif loss8g_col:
                match_8g = f"MATCH(MAX({rng_8g}),{rng_8g},0)"
                link_8g = f"\"#'{sheet_name}'!A\"&({match_8g}+2)"

                safe_formula = f'IF({v_freq}<=8, HYPERLINK({link_8g}, INDEX({net_range},{match_8g})), "No Data")'
                ws[f'L{idx}'].value = f'={safe_formula}'
                ws[f'L{idx}'].font = Font(color="0563C1", underline="single")
                ws[f'M{idx}'].value = f'=INDEX({rng_8g}, {match_8g})'
                ws[f'N{idx}'].value = "N/A"

            elif loss16g_col:
                match_16g = f"MATCH(MAX({rng_16g}),{rng_16g},0)"
                link_16g = f"\"#'{sheet_name}'!A\"&({match_16g}+2)"

                safe_formula = f'IF({v_freq}>8, HYPERLINK({link_16g}, INDEX({net_range},{match_16g})), "No Data")'
                ws[f'L{idx}'].value = f'={safe_formula}'
                ws[f'L{idx}'].font = Font(color="0563C1", underline="single")
                ws[f'M{idx}'].value = "N/A"
                ws[f'N{idx}'].value = f'=INDEX({rng_16g}, {match_16g})'

            # number format
            ws[f'H{idx}'].number_format = '0.00'
            ws[f'I{idx}'].number_format = '0.00'
            ws[f'M{idx}'].number_format = '0.00'
            ws[f'N{idx}'].number_format = '0.00'

            # Risk Level (O)
            v_spec = f"VLOOKUP($A{idx},{config_range},3,0)"
            v_over = f"VLOOKUP($A{idx},{config_range},4,0)"
            limit_val = f"({v_spec}-{v_over})"

            m_ref = f"$M{idx}"
            n_ref = f"$N{idx}"

            risk_formula = (
                f'=IFERROR('
                f'IF({v_freq}<=8,'
                f'  IF(ISNUMBER({m_ref}), IF({m_ref}>{limit_val}, "High Risk", "Low Risk"), "No Data"),'
                f'  IF(ISNUMBER({n_ref}), IF({n_ref}>{limit_val}, "High Risk", "Low Risk"), "No Data")'
                f'), "Data Error")'
            )

            ws[f'O{idx}'].value = risk_formula
            ws[f'O{idx}'].font = Font(bold=True)

    ws.add_data_validation(dv_interface)

    # 自動調整欄寬並添加邊框
    auto_adjust_column_width(ws)
    add_borders_to_sheet(ws)


def run_processing(file_board1, file_board2, file_mapping, output_path, log_func, finish_callback):
    try:
        comp_map, pin_map = load_mapping_table(file_mapping, log_func)
        df_board1 = load_all_sheets(file_board1, log_func)
        df_board2 = load_all_sheets(file_board2, log_func)

        log_func(f"Total Board1 records: {len(df_board1)}")
        log_func(f"Total Board2 records: {len(df_board2)}")

        log_func("Building index...")
        b2_start_idx = defaultdict(list)
        b2_end_idx = defaultdict(list)

        for _, row in df_board2.iterrows():
            s_pin_str = str(row.get('Start Pin', '')).strip()
            if s_pin_str and s_pin_str != 'nan':
                b2_start_idx[s_pin_str].append(row)

            e_pin_str = str(row.get('End Pin', '')).strip()
            if e_pin_str and e_pin_str != 'nan':
                b2_end_idx[e_pin_str].append(row)

        log_func("Starting pin matching...")
        merged_by_mapping = defaultdict(list)

        for _, row_b1 in df_board1.iterrows():
            def find_targets(pin_full_str):
                targets = []
                if pin_full_str in pin_map:
                    target_full = pin_map[pin_full_str]
                    src_comp, _ = parse_pin(pin_full_str)
                    targets.append((target_full, src_comp))
                comp, pin = parse_pin(pin_full_str)
                if comp and comp in comp_map:
                    for target_comp in comp_map[comp]:
                        target_full = f"{target_comp}.{pin}"
                        targets.append((target_full, comp))
                return targets

            b1_s_str = str(row_b1.get('Start Pin', '')).strip()
            b1_s_targets = find_targets(b1_s_str) if b1_s_str and b1_s_str != 'nan' else []
            b1_e_str = str(row_b1.get('End Pin', '')).strip()
            b1_e_targets = find_targets(b1_e_str) if b1_e_str and b1_e_str != 'nan' else []

            matches = []

            for target_pin_str, src_comp in b1_s_targets:
                if target_pin_str in b2_start_idx:
                    for row_b2 in b2_start_idx[target_pin_str]:
                        target_comp, _ = parse_pin(target_pin_str)
                        matches.append(('Board1_Start_to_Board2_Start', src_comp, target_comp, row_b2))
                if target_pin_str in b2_end_idx:
                    for row_b2 in b2_end_idx[target_pin_str]:
                        target_comp, _ = parse_pin(target_pin_str)
                        matches.append(('Board1_Start_to_Board2_End', src_comp, target_comp, row_b2))

            for target_pin_str, src_comp in b1_e_targets:
                if target_pin_str in b2_start_idx:
                    for row_b2 in b2_start_idx[target_pin_str]:
                        target_comp, _ = parse_pin(target_pin_str)
                        matches.append(('Board1_End_to_Board2_Start', src_comp, target_comp, row_b2))
                if target_pin_str in b2_end_idx:
                    for row_b2 in b2_end_idx[target_pin_str]:
                        target_comp, _ = parse_pin(target_pin_str)
                        matches.append(('Board1_End_to_Board2_End', src_comp, target_comp, row_b2))

            if matches:
                seen = set()
                for match_type, src_comp, target_comp, row_b2 in matches:
                    row_id = id(row_b2)
                    if row_id in seen: 
                        continue
                    seen.add(row_id)

                    mapping_key = f"{src_comp}->{target_comp}"
                    item = {'Connection_Type': match_type}
                    for k, v in row_b1.items(): 
                        item[f"Board1_{k}"] = v
                    for k, v in row_b2.items(): 
                        item[f"Board2_{k}"] = v

                    merged_by_mapping[mapping_key].append(item)

        if merged_by_mapping:
            log_func(f"Found {sum(len(v) for v in merged_by_mapping.values())} connections")
            log_func("Writing file...")

            wb = Workbook()
            wb.remove(wb.active)
            summary_info_list = []

            for mapping_key in sorted(merged_by_mapping.keys()):
                rows_list = merged_by_mapping[mapping_key]
                df_result = pd.DataFrame(rows_list).dropna(axis=1, how='all')

                # 🔍 靈活搜尋 Total Length 欄位
                b1_len_col = None
                b2_len_col = None
                
                for col in df_result.columns:
                    col_lower = col.lower()
                    if 'board1' in col_lower and 'length' in col_lower and 'total' in col_lower:
                        b1_len_col = col
                        log_func(f"  Found Board1 length column: {col}")
                    elif 'board2' in col_lower and 'length' in col_lower and 'total' in col_lower:
                        b2_len_col = col
                        log_func(f"  Found Board2 length column: {col}")

                # 計算 Total Length
                if b1_len_col or b2_len_col:
                    val1 = df_result[b1_len_col].fillna(0) if b1_len_col else 0
                    val2 = df_result[b2_len_col].fillna(0) if b2_len_col else 0
                    df_result['Total Length'] = val1 + val2
                    log_func(f"  ✅ Total Length calculated")
                else:
                    log_func(f"  ⚠️ Warning: No length columns found for {mapping_key}")

                # 搜尋 Loss 欄位
                b1_loss_8g = next((c for c in df_result.columns if 'Board1' in c and 'Loss at 8G' in c), None)
                b2_loss_8g = next((c for c in df_result.columns if 'Board2' in c and 'Loss at 8G' in c), None)
                b1_loss_16g = next((c for c in df_result.columns if 'Board1' in c and 'Loss at 16G' in c), None)
                b2_loss_16g = next((c for c in df_result.columns if 'Board2' in c and 'Loss at 16G' in c), None)

                has_loss, has_loss_8g, has_loss_16g = False, False, False
                if b1_loss_8g or b2_loss_8g:
                    val1 = df_result[b1_loss_8g].fillna(0) if b1_loss_8g else 0
                    val2 = df_result[b2_loss_8g].fillna(0) if b2_loss_8g else 0
                    df_result['Total Loss at 8G'] = val1 + val2
                    has_loss = True
                    has_loss_8g = True

                if b1_loss_16g or b2_loss_16g:
                    val1 = df_result[b1_loss_16g].fillna(0) if b1_loss_16g else 0
                    val2 = df_result[b2_loss_16g].fillna(0) if b2_loss_16g else 0
                    df_result['Total Loss at 16G'] = val1 + val2
                    has_loss = True
                    has_loss_16g = True

                # 重新排列欄位順序：Total Length 放在 Total Loss 之前
                cols = list(df_result.columns)
                total_cols = [c for c in cols if c.startswith('Total')]
                other_cols = [c for c in cols if not c.startswith('Total')]
                
                # 確保順序：Total Length -> Total Loss at 8G -> Total Loss at 16G
                ordered_total_cols = []
                if 'Total Length' in total_cols:
                    ordered_total_cols.append('Total Length')
                if 'Total Loss at 8G' in total_cols:
                    ordered_total_cols.append('Total Loss at 8G')
                if 'Total Loss at 16G' in total_cols:
                    ordered_total_cols.append('Total Loss at 16G')
                
                # 重新組合：其他欄位 + Total 欄位（按順序）
                df_result = df_result[other_cols + ordered_total_cols]

                safe_name = mapping_key[:31].replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
                ws = wb.create_sheet(safe_name)

                cols = list(df_result.columns)
                ws.append(['← Back to Summary'])
                ws.append(cols)

                for r in df_result.itertuples(index=False):
                    ws.append(list(r))

                try:
                    net_col_letter = get_column_letter(next(i for i, c in enumerate(cols) if 'net_name' in c.lower()) + 1)
                except:
                    net_col_letter = 'B'

                try:
                    len_col_letter = get_column_letter(cols.index('Total Length') + 1)
                except:
                    len_col_letter = 'Z'

                total_loss8g_col_letter = None
                total_loss16g_col_letter = None

                if has_loss_8g and 'Total Loss at 8G' in cols:
                    try:
                        total_loss8g_col_letter = get_column_letter(cols.index('Total Loss at 8G') + 1)
                    except:
                        pass

                if has_loss_16g and 'Total Loss at 16G' in cols:
                    try:
                        total_loss16g_col_letter = get_column_letter(cols.index('Total Loss at 16G') + 1)
                    except:
                        pass

                ref_parts = mapping_key.split('->')

                summary_info_list.append({
                    'sheet_name': safe_name,
                    'ref1': ref_parts[0] if len(ref_parts) > 0 else '',
                    'ref2': ref_parts[1] if len(ref_parts) > 1 else '',
                    'net_count': len(rows_list),
                    'max_row': len(rows_list) + 2,
                    'net_name_col': net_col_letter,
                    'total_len_col': len_col_letter,
                    'has_loss': has_loss,
                    'total_loss8g_col': total_loss8g_col_letter,
                    'total_loss16g_col': total_loss16g_col_letter
                })

                ws['A1'].hyperlink = '#Summary!A1'
                ws['A1'].font = Font(underline='single', color='0563C1', bold=True, size=11)
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                ws.merge_cells('A1:B1')
                ws['A1'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

                b1_font = Font(color="006400")
                b2_font = Font(color="00008B")
                total_font = Font(color="FF0000", bold=True)

                for col_idx, col_name in enumerate(cols, 1):
                    if col_name.startswith('Board1_'):
                        for row in range(3, ws.max_row + 1):
                            ws.cell(row, col_idx).font = b1_font
                    elif col_name.startswith('Board2_'):
                        for row in range(3, ws.max_row + 1):
                            ws.cell(row, col_idx).font = b2_font
                    elif col_name.startswith('Total'):
                        for row in range(3, ws.max_row + 1):
                            ws.cell(row, col_idx).font = total_font

                # 自動調整欄寬並添加邊框
                auto_adjust_column_width(ws)
                add_borders_to_sheet(ws)

                log_func(f"  Sheet '{safe_name}': {len(rows_list)} records")

            log_func("Generating Summary and Config sheets...")
            create_summary_sheet(wb, summary_info_list)
            create_config_sheet(wb)

            wb.save(output_path)
            finish_callback(True, f"Success!\n{output_path}\n\nPage 1: Summary\nPage 2: Config\nPage 3+: Data ({len(merged_by_mapping)} sheets)")
        else:
            log_func("Matching complete. No data found.")
            finish_callback(False, "No matching data found.")

    except Exception as e:
        log_func(f"Error: {e}")
        import traceback
        traceback.print_exc()
        finish_callback(False, str(e))


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Board1 & Board2 Pin Matching Tool")
        self.geometry("600x550")
        self.board1 = tk.StringVar()
        self.board2 = tk.StringVar()
        self.mapping = tk.StringVar()
        self.setup_ui()

    def setup_ui(self):
        tk.Label(self, text="Board Pin Matching Tool", font=('Arial', 14, 'bold')).pack(pady=10)
        f = tk.Frame(self)
        f.pack(fill='x', padx=20)

        self.add_file_row(f, "Board1 File:", self.board1)
        self.add_file_row(f, "Board2 File:", self.board2)
        self.add_file_row(f, "Mapping Table:", self.mapping)

        tk.Button(self, text="Start", command=self.run, bg='#4CAF50', fg='white', font=('Arial', 12)).pack(pady=15, fill='x', padx=40)

        self.log_area = scrolledtext.ScrolledText(self, height=15)
        self.log_area.pack(fill='both', expand=True, padx=20, pady=10)

    def add_file_row(self, parent, label, var):
        row = tk.Frame(parent)
        row.pack(fill='x', pady=5)
        tk.Label(row, text=label, width=12, anchor='w').pack(side='left')
        tk.Entry(row, textvariable=var).pack(side='left', fill='x', expand=True, padx=5)
        tk.Button(row, text="Browse", command=lambda: self.browse(var)).pack(side='right')

    def browse(self, var):
        try:
            f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
            if f:
                var.set(f)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open file: {e}")

    def log(self, s):
        self.log_area.insert(tk.END, s + "\n")
        self.log_area.see(tk.END)

    def run(self):
        if not (self.board1.get() and self.board2.get() and self.mapping.get()):
            messagebox.showwarning("Alert", "Please select all files")
            return

        out = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile="Board_Merged_Report.xlsx")
        if not out:
            return

        self.log_area.delete(1.0, tk.END)
        self.log("Processing...")
        threading.Thread(target=run_processing, args=(self.board1.get(), self.board2.get(), self.mapping.get(), out, self.log, self.done)).start()

    def done(self, success, msg):
        if success:
            messagebox.showinfo("Success", msg)
        else:
            messagebox.showerror("Error", msg)


if __name__ == "__main__":
    App().mainloop()
